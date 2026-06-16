"""
Report export views — PDF and Excel (CSV fallback if openpyxl missing).
"""
import csv
import io
from datetime import timedelta
from django.http import HttpResponse
from django.utils import timezone
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from inventory.models import InventoryItem, StockMovement, Product
from warehouses.models import Bin


def _get_inventory_rows():
    rows = []
    for item in InventoryItem.objects.select_related('product', 'bin__rack__zone__warehouse'):
        rows.append([
            item.product.sku,
            item.product.name,
            item.bin.code,
            item.bin.rack.name,
            item.bin.rack.zone.name,
            item.bin.rack.zone.warehouse.name,
            item.quantity,
            item.reserved_quantity,
            item.available_quantity,
            item.status,
            str(item.expiry_date) if item.expiry_date else '',
        ])
    return rows


def _get_movement_rows(days=30):
    since = timezone.now() - timedelta(days=days)
    rows = []
    for m in StockMovement.objects.filter(timestamp__gte=since).select_related(
        'product', 'bin', 'performed_by'
    ).order_by('-timestamp'):
        rows.append([
            m.timestamp.strftime('%Y-%m-%d %H:%M'),
            m.product.sku,
            m.product.name,
            m.movement_type,
            m.quantity,
            m.bin.code,
            m.performed_by.username if m.performed_by else 'System',
            m.note or '',
        ])
    return rows


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_inventory_csv(request):
    """GET /api/v1/reports/inventory/csv/"""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="inventory_report.csv"'

    writer = csv.writer(response)
    writer.writerow(['SKU', 'Product', 'Bin', 'Rack', 'Zone', 'Warehouse',
                     'Qty', 'Reserved', 'Available', 'Status', 'Expiry Date'])
    for row in _get_inventory_rows():
        writer.writerow(row)

    return response


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_movements_csv(request):
    """GET /api/v1/reports/movements/csv/?days=30"""
    days = int(request.query_params.get('days', 30))
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="movements_{days}days.csv"'

    writer = csv.writer(response)
    writer.writerow(['Timestamp', 'SKU', 'Product', 'Type', 'Qty', 'Bin', 'By', 'Note'])
    for row in _get_movement_rows(days=days):
        writer.writerow(row)

    return response


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_inventory_excel(request):
    """GET /api/v1/reports/inventory/excel/  — requires openpyxl"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return export_inventory_csv(request)  # Fallback to CSV

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventory Report"

    headers = ['SKU', 'Product', 'Bin', 'Rack', 'Zone', 'Warehouse',
               'Qty', 'Reserved', 'Available', 'Status', 'Expiry Date']

    header_fill = PatternFill("solid", fgColor="4F46E5")
    header_font = Font(bold=True, color="FFFFFF")

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    for row_idx, row in enumerate(_get_inventory_rows(), start=2):
        for col_idx, val in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=val)

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="inventory_report.xlsx"'
    return response


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def warehouse_utilization_report(request):
    """GET /api/v1/reports/utilization/  — JSON summary for analytics chart"""
    from rest_framework.response import Response
    from warehouses.models import Zone

    zones = []
    for zone in Zone.objects.prefetch_related('racks__bins__inventory_items'):
        total_capacity = 0
        total_used = 0
        for rack in zone.racks.all():
            for bin_obj in rack.bins.all():
                total_capacity += bin_obj.capacity
                total_used += sum(i.quantity for i in bin_obj.inventory_items.all())

        zones.append({
            'zone': zone.name,
            'warehouse': zone.warehouse.name,
            'total_capacity': total_capacity,
            'used': total_used,
            'available': max(0, total_capacity - total_used),
            'percent': round((total_used / total_capacity * 100) if total_capacity > 0 else 0, 1),
        })

    return Response({'zones': zones, 'generated_at': timezone.now()})


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def movement_trends(request):
    """GET /api/v1/reports/movement-trends/?days=30"""
    from rest_framework.response import Response
    from collections import defaultdict

    days = int(request.query_params.get('days', 30))
    since = timezone.now() - timedelta(days=days)

    daily = defaultdict(lambda: {'in': 0, 'out': 0})
    for m in StockMovement.objects.filter(timestamp__gte=since):
        date_key = m.timestamp.strftime('%m/%d')

        if m.movement_type in ('in', 'receiving'):
            daily[date_key]['in'] += m.quantity
        elif m.movement_type in ('out', 'retrieval'):
            daily[date_key]['out'] += m.quantity

    trends = [{'date': k, 'in': v['in'], 'out': v['out']} for k, v in sorted(daily.items())]
    return Response({'trends': trends})

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def low_stock_analytics(request):
    """GET /api/v1/inventory/low-stock/  — for Analytics Dashboard"""
    from rest_framework.response import Response
    from django.db.models import Sum
    results = []
    for product in Product.objects.all():
        total_qty = InventoryItem.objects.filter(product=product).aggregate(
            total=Sum('quantity'))['total'] or 0
        if total_qty <= product.reorder_level:
            results.append({
                'sku': product.sku,
                'name': product.name,
                'quantity': total_qty,
                'reorder_level': product.reorder_level,
            })
    return Response(results)